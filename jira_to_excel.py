import urllib.request
import base64
import json
import os
import sys
from datetime import datetime, date
from collections import Counter

# Tum sirlar ve sabit degerler config.py'den okunur (DRY + guvenlik).
sys.path.insert(0, "/opt/jira_rapor")
import config as cfg

TOKEN    = cfg.JIRA_TOKEN
EMAIL    = cfg.JIRA_EMAIL
BASE_URL = cfg.JIRA_BASE
START_DATE_FIELD = cfg.JIRA_START_DATE_FIELD
TODAY = date.today()

credentials = base64.b64encode(f"{EMAIL}:{TOKEN}".encode()).decode()
req_headers = {
    "Authorization": f"Basic {credentials}",
    "Content-Type": "application/json",
    "Accept": "application/json"
}

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

# ── helpers ────────────────────────────────────────────────────────────────

def fetch_issues(project_key):
    fields = ["summary", "status", "assignee", "priority", "issuetype",
              "created", "duedate", "reporter", "labels",
              "subtasks", "parent", "description", START_DATE_FIELD]
    issues_raw = []
    next_token = None
    while True:
        body = {
            "jql": f"project={project_key} ORDER BY key ASC",
            "maxResults": 500,
            "fields": fields,
        }
        if next_token:
            body["nextPageToken"] = next_token
        payload = json.dumps(body).encode("utf-8")
        req = urllib.request.Request(
            f"{BASE_URL}/rest/api/3/search/jql",
            data=payload, headers=req_headers, method="POST"
        )
        with urllib.request.urlopen(req) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        issues_raw.extend(data.get("issues", []))
        if data.get("isLast", True):
            break
        next_token = data.get("nextPageToken")
        if not next_token:
            break

    # Build ordered list: parent immediately followed by its subtasks (board order)
    issue_map = {i["key"]: i for i in issues_raw}
    children_map = {}
    # 1) Classic subtasks: parent's "subtasks" field
    for i in issues_raw:
        subs = i["fields"].get("subtasks", [])
        if subs:
            children_map[i["key"]] = [s["key"] for s in subs]
    # 2) Next-gen child issues: child's "parent" field (may not appear in
    #    parent's subtasks list, so we must also scan from the child side)
    for i in issues_raw:
        parent_info = i["fields"].get("parent")
        if parent_info:
            pk = parent_info["key"]
            children_map.setdefault(pk, [])
            if i["key"] not in children_map[pk]:
                children_map[pk].append(i["key"])

    ordered = []
    for i in issues_raw:
        key = i["key"]
        if not (i["fields"].get("parent") or i["fields"]["issuetype"].get("subtask")):
            ordered.append(key)
            for ck in children_map.get(key, []):
                ordered.append(ck)

    return [issue_map[k] for k in ordered if k in issue_map]

def parse_date(dt_str):
    if not dt_str:
        return None
    try:
        return date.fromisoformat(dt_str[:10])
    except Exception:
        return None

TR_MONTHS = {
    1: "Ocak", 2: "Şubat", 3: "Mart", 4: "Nisan",
    5: "Mayıs", 6: "Haziran", 7: "Temmuz", 8: "Ağustos",
    9: "Eylül", 10: "Ekim", 11: "Kasım", 12: "Aralık"
}

def format_tr_date(d):
    if not d:
        return ""
    return f"{d.day:02d}.{d.month:02d}.{d.year}"

def days_remaining(due_date):
    return (due_date - TODAY).days if due_date else None

def extract_adf_text(node):
    if not node:
        return ""
    if isinstance(node, str):
        return node
    if isinstance(node, dict):
        if node.get("type") == "text":
            return node.get("text", "")
        parts = []
        for child in node.get("content", []):
            t = extract_adf_text(child)
            if t:
                parts.append(t)
        sep = "\n" if node.get("type") in ("doc", "paragraph", "bulletList", "listItem", "orderedList") else " "
        return sep.join(parts)
    return ""

def build_rows(issues):
    # First pass: collect due dates by key
    due_by_key = {}
    for issue in issues:
        d = parse_date(issue["fields"].get("duedate"))
        due_by_key[issue["key"]] = d

    rows = []
    for issue in issues:
        f = issue["fields"]
        is_sub = bool(f.get("parent") or f["issuetype"].get("subtask"))
        is_done = (f["status"].get("statusCategory") or {}).get("key") == "done"

        # Erken bitirme tespiti: durum done ve max(due, start) > TODAY ise erken
        # Tip 1: bitis tarihi gelecekte (due > today) - bitisten once bitirildi
        # Tip 2: baslangic tarihi gelecekte (start > today) - baslamadan once bitirildi
        early_days = None
        late_start_days = None  # Tip 2 geciken: start gecmis ama hala baslanmamis
        if is_done:
            own_due = due_by_key.get(issue["key"])
            start_dt = parse_date(f.get(START_DATE_FIELD))
            candidates = [d for d in (own_due, start_dt) if d]
            if candidates:
                target = max(candidates)
                if target > TODAY:
                    early_days = (target - TODAY).days
        else:
            # Done degil: baslangic gecmis ama henuz baslanmamis mi?
            status_cat = (f["status"].get("statusCategory") or {}).get("key", "")
            start_dt = parse_date(f.get(START_DATE_FIELD))
            not_started = (status_cat == "new")
            if not_started and start_dt and start_dt < TODAY:
                late_start_days = (TODAY - start_dt).days

        due_date = parse_date(f.get("duedate"))
        remaining = days_remaining(due_date)
        desc_text = ""
        if f["status"]["name"] == "Beklemede":
            desc_text = extract_adf_text(f.get("description")).strip()

        # Date conflict check: subtask due date > parent due date
        date_conflict = False
        if is_sub and due_date:
            parent_key = (f.get("parent") or {}).get("key")
            if parent_key:
                parent_due = due_by_key.get(parent_key)
                if parent_due and due_date > parent_due:
                    date_conflict = True

        rows.append({
            "Anahtar":           issue["key"],
            "Konu Türü":         f["issuetype"]["name"],
            "Özet":              f["summary"],
            "Durum":             f["status"]["name"],
            "Öncelik":           (f.get("priority") or {}).get("name", ""),
            "Atanan":            (f.get("assignee") or {}).get("displayName", ""),
            "Raporlayan":        (f.get("reporter") or {}).get("displayName", ""),
            "Etiketler":         ", ".join(f.get("labels") or []),
            "Oluşturma Tarihi":  format_tr_date(parse_date(f.get("created"))),
            "Başlangıç Tarihi":  format_tr_date(parse_date(f.get(START_DATE_FIELD))),
            "Son Tarih":         format_tr_date(due_date),
            "Kalan Gün":         remaining,
            "İlerleme":          remaining,
            "Bekleme Açıklaması": desc_text,
            "_subtask":          is_sub,
            "_date_conflict":    date_conflict,
            "_done":             is_done,
            "_early_days":       early_days,
            "_late_start_days":  late_start_days,
        })
    return rows

# ── Excel writing ───────────────────────────────────────────────────────────

DISPLAY_COLS = [
    "Anahtar", "Konu Türü", "Özet", "Durum", "Öncelik",
    "Atanan", "Raporlayan", "Etiketler",
    "Oluşturma Tarihi", "Başlangıç Tarihi", "Son Tarih", "Kalan Gün", "İlerleme",
    "Bekleme Açıklaması"
]
COL_WIDTHS = {
    "Anahtar": 13, "Konu Türü": 16, "Özet": 52, "Durum": 16,
    "Öncelik": 12, "Atanan": 22, "Raporlayan": 22, "Etiketler": 20,
    "Oluşturma Tarihi": 18, "Başlangıç Tarihi": 18, "Son Tarih": 16,
    "Kalan Gün": 14, "İlerleme": 30, "Bekleme Açıklaması": 48
}
HEADER_COMMENTS = {
    "Anahtar":           "Jira'daki benzersiz görev kodu.\nAna görevler kalın mavi, alt görevler gri ve girintili gösterilir.",
    "Konu Türü":         "Görevin türü: Epic, Story, Task, Sub-task, Hata vb.",
    "Özet":              "Görevin başlığı / kısa açıklaması.\nAna görevler kalın, alt görevler italik gösterilir.",
    "Durum":             "Görevin mevcut durumu:\n• Yapılacaklar — henüz başlanmadı\n• Devam Ediyor — üzerinde çalışılıyor\n• Beklemede — bir engel/beklenti var\n• Tamamlandı — teslim edildi",
    "Öncelik":           "Görevin öncelik seviyesi:\nHighest > High > Medium > Low > Lowest",
    "Atanan":            "Görevi üstlenen kişi.",
    "Raporlayan":        "Görevi oluşturan / talep eden kişi.",
    "Etiketler":         "Göreve eklenen serbest etiketler.\nFiltreleme ve gruplama için kullanılır.",
    "Oluşturma Tarihi":  "Görevin Jira'da ilk oluşturulduğu tarih.",
    "Başlangıç Tarihi":  "Görevin çalışmaya başlanacağı planlanan tarih (Start Date).",
    "Son Tarih":         "Görevin tamamlanması gereken bitiş tarihi.",
    "Kalan Gün":         "Bugünden bitiş tarihine kaç gün kaldığı.\n🔴 Kırmızı: süre geçmiş veya ≤3 gün\n🟠 Turuncu: ≤7 gün\n🟢 Yeşil: 14+ gün",
    "İlerleme":          "Bitiş tarihine göre görsel ilerleme çubuğu.\n█ dolu = geçen süre  ░ boş = kalan süre\nReferans penceresi: 30 gün.",
    "Bekleme Açıklaması":"Yalnızca 'Beklemede' durumundaki görevler için.\nJira açıklaması buraya taşınır."
}
STATUS_COLORS = {
    "Yapılacaklar": "FFF2CC",
    "Devam Ediyor":  "DEEAF1",
    "Tamamlandı":    "E2EFDA",
    "Beklemede":     "FCE4D6",
    "İptal Edildi":  "F2F2F2"
}
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def write_sheet(wb, sheet_title, rows, header_color="1F4E79"):
    ws = wb.create_sheet(sheet_title)
    header_fill = PatternFill("solid", fgColor=header_color)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(DISPLAY_COLS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_name, 15)
        if col_name in HEADER_COMMENTS:
            c = Comment(HEADER_COMMENTS[col_name], "Jira Raporu")
            c.width = 280
            c.height = max(80, HEADER_COMMENTS[col_name].count("\n") * 18 + 40)
            cell.comment = c

    ws.row_dimensions[1].height = 30

    PROG_COL = DISPLAY_COLS.index("İlerleme") + 1

    for row_idx, row_data in enumerate(rows, 2):
        status = row_data["Durum"]
        is_sub = row_data["_subtask"]
        is_done = row_data.get("_done", False)
        early_days = row_data.get("_early_days")
        late_start_days = row_data.get("_late_start_days")
        conflict = row_data.get("_date_conflict", False)
        row_color = STATUS_COLORS.get(status, "FFFFFF")
        fill = PatternFill("solid", fgColor=row_color)
        conflict_fill = PatternFill("solid", fgColor="FF0000")
        conflict_fill_light = PatternFill("solid", fgColor="FFB3B3")
        remaining = row_data["İlerleme"]

        for col_idx, col_name in enumerate(DISPLAY_COLS, 1):
            val = row_data[col_name]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(col_name == "Özet"))

            if col_name == "İlerleme":
                cell.value = ""
                cell.fill = conflict_fill if conflict else PatternFill("solid", fgColor="F2F2F2")
                if conflict:
                    cell.value = "⚠ TARİHİ DÜZELT!"
                    cell.font = Font(bold=True, color="FFFFFF", size=10)
                    cell.fill = conflict_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                continue

            cell.fill = conflict_fill_light if conflict else fill

            if col_name == "Anahtar":
                if conflict:
                    cell.font = Font(bold=True, color="CC0000", size=10)
                    cell.value = "  └ ⚠ " + str(val)
                elif is_sub:
                    cell.font = Font(italic=True, color="666666", size=10)
                    cell.value = "  └ " + str(val)
                else:
                    cell.font = Font(bold=True, color="1F4E79", size=11)

            elif col_name == "Özet":
                if conflict:
                    cell.font = Font(bold=True, color="CC0000", size=10)
                    cell.value = "    " + str(val) + "  ← TARİHİ DÜZELT"
                elif is_sub:
                    cell.font = Font(italic=True, color="555555", size=10)
                    cell.value = "    " + str(val)
                else:
                    cell.font = Font(bold=True, color="000000", size=11)

            elif col_name == "Kalan Gün":
                if is_done:
                    if early_days:
                        cell.value = f"{early_days} gün ERKEN bitirildi"
                        cell.font = Font(bold=True, color="148F3A")
                        cell.fill = PatternFill("solid", fgColor="D4EFDF")
                    else:
                        cell.value = "✓ Tamamlandı"
                        cell.font = Font(bold=True, color="148F3A")
                        cell.fill = PatternFill("solid", fgColor="E2EFDA")
                elif late_start_days:
                    cell.value = f"Başlangıç gecikti ({late_start_days} g)"
                    cell.font = Font(bold=True, color="D35400")
                    cell.fill = PatternFill("solid", fgColor="FFE5CC")
                elif remaining is None:
                    cell.value = "—"
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif remaining < 0:
                    cell.value = f"Geçti ({abs(remaining)} gün)"
                    cell.font = Font(bold=True, color="FF0000")
                    cell.fill = PatternFill("solid", fgColor="FFE0E0")
                elif remaining == 0:
                    cell.value = "Bugün!"
                    cell.font = Font(bold=True, color="FF6600")
                    cell.fill = PatternFill("solid", fgColor="FFEEDD")
                elif remaining <= 3:
                    cell.value = f"{remaining} gün"
                    cell.font = Font(bold=True, color="FF6600")
                    cell.fill = PatternFill("solid", fgColor="FFEEDD")
                elif remaining <= 7:
                    cell.value = f"{remaining} gün"
                    cell.font = Font(color="CC8800")
                else:
                    cell.value = f"{remaining} gün"
                    cell.font = Font(color="228B22")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            elif col_name == "Son Tarih" and conflict:
                cell.font = Font(bold=True, color="FFFFFF", size=10)
                cell.fill = conflict_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            elif col_name == "Bekleme Açıklaması":
                if conflict:
                    cell.value = "⚠ Ana görevin bitiş tarihi bu alt görevden önce! Tarihi düzeltin."
                    cell.font = Font(bold=True, color="CC0000", size=10)
                    cell.fill = conflict_fill_light
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                elif val:
                    cell.font = Font(italic=True, color="7B4F00", size=10)
                    cell.fill = PatternFill("solid", fgColor="FFF8E7")
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                else:
                    cell.value = ""
                    cell.fill = conflict_fill_light if conflict else fill

        # Progress bar cell
        pc = ws.cell(row=row_idx, column=PROG_COL)
        pc.border = BORDER
        pc.alignment = Alignment(horizontal="left", vertical="center")
        if is_done:
            if early_days:
                pc.value = f"██████████  ✓ {early_days} GÜN ERKEN BİTİRİLDİ"
                pc.font = Font(bold=True, color="148F3A")
                pc.fill = PatternFill("solid", fgColor="D4EFDF")
            else:
                pc.value = "██████████  ✓ TAMAMLANDI"
                pc.font = Font(bold=True, color="148F3A")
                pc.fill = PatternFill("solid", fgColor="E2EFDA")
        elif late_start_days:
            pc.value = f"⚠ BAŞLANGIÇ GECİKTİ ({late_start_days} GÜN)"
            pc.font = Font(bold=True, color="D35400")
            pc.fill = PatternFill("solid", fgColor="FFE5CC")
        elif remaining is None:
            pc.value = "Tarih yok"
            pc.font = Font(color="AAAAAA", italic=True)
            pc.fill = PatternFill("solid", fgColor="F5F5F5")
        elif remaining < 0:
            pc.value = "██████████  SÜRE DOLDU"
            pc.font = Font(bold=True, color="CC0000")
            pc.fill = PatternFill("solid", fgColor="FFD0D0")
        elif remaining == 0:
            pc.value = "██████████  BUGÜN SON GÜN!"
            pc.font = Font(bold=True, color="FF6600")
            pc.fill = PatternFill("solid", fgColor="FFEEDD")
        else:
            MAX_DAYS = cfg.PROGRESS_WINDOW_PANO
            ratio = min(max(0, MAX_DAYS - remaining) / MAX_DAYS, 1.0)
            filled = round(ratio * 10)
            bar = "█" * filled + "░" * (10 - filled)
            if remaining > 14:
                bar_color, bg = "228B22", "E8F5E9"
            elif remaining > 7:
                bar_color, bg = "CC8800", "FFF9E6"
            elif remaining > 3:
                bar_color, bg = "FF6600", "FFF0E0"
            else:
                bar_color, bg = "CC0000", "FFE8E8"
            pc.value = f"{bar}  {remaining} gün kaldı"
            pc.font = Font(color=bar_color, bold=(remaining <= 7))
            pc.fill = PatternFill("solid", fgColor=bg)

        ws.row_dimensions[row_idx].height = 20 if not is_sub else 18

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    return ws

def write_summary(wb, title, rows, header_color):
    ws = wb.create_sheet(title)
    ws["A1"] = title.replace(" Özet", " — Özet")
    ws["A1"].font = Font(bold=True, size=14, color=header_color)
    ws["A2"] = f"Rapor Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A3"] = f"Toplam Issue: {len(rows)}"
    ws["A4"] = f"Ana Görev: {sum(1 for r in rows if not r['_subtask'])}   Alt Görev: {sum(1 for r in rows if r['_subtask'])}"

    status_counts = Counter(r["Durum"] for r in rows)
    ws["A6"] = "Duruma Göre Dağılım"
    ws["A6"].font = Font(bold=True)
    for i, (s, c) in enumerate(sorted(status_counts.items()), 7):
        ws[f"A{i}"] = s
        ws[f"B{i}"] = c

    off = len(status_counts) + 9
    ws[f"A{off}"] = "Önceliğe Göre Dağılım"
    ws[f"A{off}"].font = Font(bold=True)
    pri_counts = Counter(r["Öncelik"] for r in rows if r["Öncelik"])
    for i, (p, c) in enumerate(sorted(pri_counts.items()), off + 1):
        ws[f"A{i}"] = p
        ws[f"B{i}"] = c

    off2 = off + len(pri_counts) + 3
    ws[f"A{off2}"] = "Süre Durumu"
    ws[f"A{off2}"].font = Font(bold=True)
    entries = [
        ("Süresi Geçmiş",    sum(1 for r in rows if isinstance(r["İlerleme"], int) and r["İlerleme"] < 0),  "CC0000"),
        ("Bugün Son Gün",    sum(1 for r in rows if isinstance(r["İlerleme"], int) and r["İlerleme"] == 0),  "FF6600"),
        ("Bu Hafta Bitiyor", sum(1 for r in rows if isinstance(r["İlerleme"], int) and 1 <= r["İlerleme"] <= 7), "CC8800"),
        ("Tarih Girilmemiş", sum(1 for r in rows if r["İlerleme"] is None), "888888"),
    ]
    for i, (label, count, color) in enumerate(entries, off2 + 1):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = count
        ws[f"B{i}"].font = Font(color=color, bold=True)

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 10

# ── Main ────────────────────────────────────────────────────────────────────

# Projeler config.PROJECTS'den, temalar config.PROJECT_THEMES'den.
# Yeni proje eklemek icin sadece config.py'i guncelle.
projects = [
    (p, f"{p} Panosu", cfg.PROJECT_THEMES[p]["sheet_color"])
    for p in cfg.PROJECTS
]

wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

all_project_rows = {}
for proj_key, sheet_name, color in projects:
    print(f"{proj_key} verileri çekiliyor...")
    issues = fetch_issues(proj_key)
    print(f"  {len(issues)} issue bulundu.")
    rows = build_rows(issues)
    all_project_rows[sheet_name] = (rows, color)
    write_sheet(wb, sheet_name, rows, header_color=color)
    write_summary(wb, f"{sheet_name} Özet", rows, header_color=color)

# ── Ana Sayfa ──────────────────────────────────────────────────────────────
ws_home = wb.create_sheet("Ana Sayfa")
wb.move_sheet("Ana Sayfa", offset=-len(wb.sheetnames) + 1)

# ---- Card layout (dinamik: kart sayisi/genisligi degistirilebilir) ------
# Kart konfigurasyonu config.PROJECT_THEMES'den uretilir.
proj_cfg = {
    f"{p} Panosu": {
        "label":       cfg.PROJECT_THEMES[p]["label"],
        "accent":      cfg.PROJECT_THEMES[p]["card_accent"],
        "btn":         cfg.PROJECT_THEMES[p]["card_btn"],
        "stat_bg":     cfg.PROJECT_THEMES[p]["card_stat_bg"],
        "stat_border": cfg.PROJECT_THEMES[p]["card_stat_border"],
    }
    for p in cfg.PROJECTS
}

CARD_ROW_START = 8
CARD_WIDTH = 6                       # her kartin sutun sayisi
CARD_GAP = 1                         # kartlar arasi bosluk sutunu
NUM_CARDS = len(all_project_rows)
# Kart baslangic sutunlari: 3, 3+W+G, 3+2*(W+G), ...
CARD_COL_STARTS = [3 + i * (CARD_WIDTH + CARD_GAP) for i in range(NUM_CARDS)]
LAST_COL = CARD_COL_STARTS[-1] + CARD_WIDTH - 1   # son kartin son sutunu
FIRST_COL = CARD_COL_STARTS[0]                    # ilk kartin ilk sutunu = 3

BG = "F7F9FC"
# Arkaplan: tum kart alanini + 1 sag bosluk kapsa
for r in range(1, 80):
    ws_home.row_dimensions[r].height = 16
    for col in range(1, LAST_COL + 2):
        ws_home.cell(row=r, column=col).fill = PatternFill("solid", fgColor=BG)

# Column layout: A,B = 2,2 (sol bosluk); C..LAST_COL = 5 (kartlar ayni genislikte)
ws_home.column_dimensions["A"].width = 2
ws_home.column_dimensions["B"].width = 2
for col in range(FIRST_COL, LAST_COL + 1):
    ws_home.column_dimensions[get_column_letter(col)].width = 5

# ---- Header banner ----
ws_home.row_dimensions[2].height = 6
ws_home.row_dimensions[3].height = 52
ws_home.row_dimensions[4].height = 6

# Header arkaplan: ilk kartin ilk sutunundan son kartin son sutununa
for col in range(FIRST_COL, LAST_COL + 1):
    ws_home.cell(row=3, column=col).fill = PatternFill("solid", fgColor="1B3A6B")

h = ws_home.cell(row=3, column=FIRST_COL, value="  GUNDOGDU GIDA  —  JIRA PROJE PANOLARI")
h.font = Font(bold=True, size=18, color="FFFFFF", name="Calibri")
h.fill = PatternFill("solid", fgColor="1B3A6B")
h.alignment = Alignment(horizontal="left", vertical="center")
ws_home.merge_cells(start_row=3, start_column=FIRST_COL, end_row=3, end_column=LAST_COL)

ws_home.row_dimensions[5].height = 22
dt = ws_home.cell(row=5, column=FIRST_COL, value=f"Rapor Tarihi:  {datetime.now().strftime('%d %B %Y  —  %H:%M')}")
dt.font = Font(size=10, color="5A6A85", italic=True, name="Calibri")
dt.alignment = Alignment(horizontal="left", vertical="center")
ws_home.merge_cells(start_row=5, start_column=FIRST_COL, end_row=5, end_column=LAST_COL)

ws_home.row_dimensions[6].height = 4
for col in range(FIRST_COL, LAST_COL + 1):
    ws_home.cell(row=6, column=col).fill = PatternFill("solid", fgColor="2E75B6")
ws_home.row_dimensions[7].height = 14

for card_idx, (sheet_name, (rows, _color)) in enumerate(all_project_rows.items()):
    # NOT: lokal degisken adi 'card' (modul 'cfg'i golgelemez)
    card = proj_cfg.get(sheet_name, {"label": sheet_name, "accent": "333333", "btn": "555555", "stat_bg": "F5F5F5", "stat_border": "CCCCCC"})
    cs = CARD_COL_STARTS[card_idx]
    ce = cs + CARD_WIDTH - 1
    r = CARD_ROW_START

    def fill_row(row_n, hex_color, height=None):
        for col in range(cs, ce + 1):
            ws_home.cell(row=row_n, column=col).fill = PatternFill("solid", fgColor=hex_color)
        if height:
            ws_home.row_dimensions[row_n].height = height

    def merged_cell(row_n, value, font_obj, fill_hex, align_h="left", align_v="center", indent=0):
        cell = ws_home.cell(row=row_n, column=cs, value=value)
        cell.font = font_obj
        cell.fill = PatternFill("solid", fgColor=fill_hex)
        cell.alignment = Alignment(horizontal=align_h, vertical=align_v, indent=indent)
        ws_home.merge_cells(start_row=row_n, start_column=cs, end_row=row_n, end_column=ce)
        return cell

    def stat_row(row_n, label, value, val_color="1B3A6B"):
        ws_home.row_dimensions[row_n].height = 20
        lc = ws_home.cell(row=row_n, column=cs, value=label)
        lc.font = Font(size=10, color="2C3E50", name="Calibri")
        lc.fill = PatternFill("solid", fgColor=card["stat_bg"])
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws_home.merge_cells(start_row=row_n, start_column=cs, end_row=row_n, end_column=ce - 2)

        vc = ws_home.cell(row=row_n, column=ce - 1, value=value)
        vc.font = Font(size=11, bold=True, color=val_color, name="Calibri")
        vc.fill = PatternFill("solid", fgColor=card["stat_bg"])
        vc.alignment = Alignment(horizontal="center", vertical="center")
        ws_home.merge_cells(start_row=row_n, start_column=ce - 1, end_row=row_n, end_column=ce)

    # Top accent bar
    fill_row(r, card["accent"], height=5); r += 1

    # Project title header
    ws_home.row_dimensions[r].height = 38
    tc = merged_cell(r, card["label"],
                     Font(bold=True, size=13, color="FFFFFF", name="Calibri"),
                     card["accent"], align_h="left", indent=1)
    # Internal hyperlink — use openpyxl Hyperlink with location
    from openpyxl.worksheet.hyperlink import Hyperlink
    tc.hyperlink = Hyperlink(ref=tc.coordinate, location=f"'{sheet_name}'!A1", tooltip=f"{sheet_name} panosuna git")
    tc.font = Font(bold=True, size=13, color="FFFFFF", name="Calibri", underline="single")
    r += 1

    # Thin divider
    fill_row(r, card["stat_border"], height=3); r += 1

    # Stats
    total      = len(rows)
    ana        = sum(1 for x in rows if not x["_subtask"])
    alt        = sum(1 for x in rows if x["_subtask"])
    tamamlandi = sum(1 for x in rows if x["Durum"] == "Tamamlandi" or "Tamamland" in x["Durum"])
    devam      = sum(1 for x in rows if x["Durum"] == "Devam Ediyor")
    bekleme    = sum(1 for x in rows if x["Durum"] == "Beklemede")
    yapilacak  = sum(1 for x in rows if x["Durum"] == "Yapilacaklar" or "Yap" in x["Durum"])
    gecmis     = sum(1 for x in rows if isinstance(x["İlerleme"], int) and x["İlerleme"] < 0)
    bu_hafta   = sum(1 for x in rows if isinstance(x["İlerleme"], int) and 0 <= x["İlerleme"] <= 7)

    stat_row(r, "Toplam Issue",       total,      "1B3A6B"); r += 1
    stat_row(r, "Ana Gorev",          ana,        "1B3A6B"); r += 1
    stat_row(r, "Alt Gorev",          alt,        "1B3A6B"); r += 1
    fill_row(r, card["stat_border"], height=3);               r += 1
    stat_row(r, "Tamamlandi",         tamamlandi, "27AE60"); r += 1
    stat_row(r, "Devam Ediyor",       devam,      "2980B9"); r += 1
    stat_row(r, "Beklemede",          bekleme,    "E67E22"); r += 1
    stat_row(r, "Yapilacaklar",       yapilacak,  "7F8C8D"); r += 1
    fill_row(r, card["stat_border"], height=3);               r += 1
    stat_row(r, "Suresi Gecmis",      gecmis,     "C0392B"); r += 1
    stat_row(r, "Bu Hafta Bitiyor",   bu_hafta,   "D35400"); r += 1
    fill_row(r, card["stat_border"], height=3);               r += 1

    # Button row
    ws_home.row_dimensions[r].height = 28
    bc = ws_home.cell(row=r, column=cs, value=">> Panoya Git  (Tiklayin)")
    bc.font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    bc.fill = PatternFill("solid", fgColor=card["btn"])
    bc.alignment = Alignment(horizontal="center", vertical="center")
    bc.hyperlink = Hyperlink(ref=bc.coordinate, location=f"'{sheet_name}'!A1", tooltip=f"{sheet_name} panosuna git")
    ws_home.merge_cells(start_row=r, start_column=cs, end_row=r, end_column=ce)
    r += 1

    # Bottom accent bar
    fill_row(r, card["accent"], height=5); r += 1

ws_home.sheet_view.showGridLines = False
ws_home.sheet_view.zoomScale = 110
ws_home.sheet_properties.tabColor = "1B3A6B"

# ── Yol Haritası (Deadline Takvimi) ────────────────────────────────────────
ws_road = wb.create_sheet("Yol Haritasi")
wb.move_sheet("Yol Haritasi", offset=-(len(wb.sheetnames) - 2))  # after Ana Sayfa

# Collect all issues with due dates from all projects
all_deadlines = []
for sheet_name, (rows, _) in all_project_rows.items():
    proj_label = sheet_name.replace(" Panosu", "")
    for r in rows:
        if r["Son Tarih"]:
            try:
                d = date(int(r["Son Tarih"][6:10]), int(r["Son Tarih"][3:5]), int(r["Son Tarih"][:2]))
                all_deadlines.append({**r, "_proj": proj_label, "_due": d})
            except Exception:
                pass

all_deadlines.sort(key=lambda x: x["_due"])

# Background
for row in range(1, 300):
    ws_road.row_dimensions[row].height = 16
    for col in range(1, 16):
        ws_road.cell(row=row, column=col).fill = PatternFill("solid", fgColor="F7F9FC")

# Column widths
ws_road.column_dimensions["A"].width = 2
ws_road.column_dimensions["B"].width = 10   # Proje
ws_road.column_dimensions["C"].width = 12   # Anahtar
ws_road.column_dimensions["D"].width = 44   # Görev Adı
ws_road.column_dimensions["E"].width = 16   # Durum
ws_road.column_dimensions["F"].width = 16   # Başlangıç Tarihi
ws_road.column_dimensions["G"].width = 16   # Son Tarih
ws_road.column_dimensions["H"].width = 14   # Kalan Gün
ws_road.column_dimensions["I"].width = 30   # Zaman Çubuğu
ws_road.column_dimensions["J"].width = 18   # Atanan

# Header banner
ws_road.row_dimensions[2].height = 6
ws_road.row_dimensions[3].height = 48
for col in range(2, 11):
    ws_road.cell(row=3, column=col).fill = PatternFill("solid", fgColor="1B3A6B")

htitle = ws_road.cell(row=3, column=2, value="  YOL HARITASI  —  TUM PROJELER DEADLINE TAKVIMI")
htitle.font = Font(bold=True, size=17, color="FFFFFF", name="Calibri")
htitle.alignment = Alignment(horizontal="left", vertical="center")
ws_road.merge_cells("B3:J3")

ws_road.row_dimensions[4].height = 6
for col in range(2, 11):
    ws_road.cell(row=4, column=col).fill = PatternFill("solid", fgColor="2E75B6")

ws_road.row_dimensions[5].height = 22
info = ws_road.cell(row=5, column=2,
    value="Bugün: {}   |   Tarihli görev sayısı: {}   |   {}".format(
        TODAY.strftime('%d.%m.%Y'), len(all_deadlines),
        "   ".join("{}: {}".format(p[0], sum(1 for x in all_deadlines if x["_proj"]==p[0])) for p in projects)))
info.font = Font(size=10, color="5A6A85", italic=True, name="Calibri")
info.alignment = Alignment(horizontal="left", vertical="center")
ws_road.merge_cells("B5:J5")

ws_road.row_dimensions[6].height = 5
for col in range(2, 11):
    ws_road.cell(row=6, column=col).fill = PatternFill("solid", fgColor="DDDDDD")

# Column headers
ws_road.row_dimensions[7].height = 24
HDR_FILL = PatternFill("solid", fgColor="2C3E50")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
HDR_ALIGN = Alignment(horizontal="center", vertical="center")
for col, label in zip(
    ["B","C","D","E","F","G","H","I","J"],
    ["Proje","Anahtar","Görev Adı","Durum","Başlangıç Tarihi","Son Tarih","Kalan Gün","Zaman Çubuğu","Atanan"]
):
    c = ws_road[f"{col}7"]
    c.value = label
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = HDR_ALIGN
    c.border = BORDER

ws_road.freeze_panes = "B8"

# Month group colors (alternating)
MONTH_FILLS = ["EBF2FA", "E8F5E9", "FFF8E7", "F3E8FF", "FFE8E8", "E8F8FF"]
THIN2 = Side(style="thin", color="CCCCCC")
BORDER2 = Border(left=THIN2, right=THIN2, top=THIN2, bottom=THIN2)

current_month = None
month_fill_idx = 0
row_n = 8

for item in all_deadlines:
    d = item["_due"]
    proj = item["_proj"]
    remaining = item["İlerleme"]
    is_sub = item["_subtask"]
    status = item["Durum"]
    is_done = item.get("_done", False)
    early_days = item.get("_early_days")
    late_start_days = item.get("_late_start_days")

    # Month separator
    month_key = (d.year, d.month)
    if month_key != current_month:
        current_month = month_key
        month_fill_idx = (month_fill_idx + 1) % len(MONTH_FILLS)

        ws_road.row_dimensions[row_n].height = 22
        ay_label = f"  {TR_MONTHS[d.month].upper()}  {d.year}"
        mc = ws_road.cell(row=row_n, column=2, value=ay_label)
        mc.font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
        mc.fill = PatternFill("solid", fgColor="34495E")
        mc.alignment = Alignment(horizontal="left", vertical="center")
        ws_road.merge_cells(start_row=row_n, start_column=2, end_row=row_n, end_column=10)
        row_n += 1

    # Row background
    if is_done:
        row_bg = "E8F5E9"
    elif late_start_days:
        row_bg = "FFE5CC"
    elif remaining is not None and remaining < 0:
        row_bg = "FFE8E8"
    elif remaining is not None and remaining <= 7:
        row_bg = "FFF4E0"
    else:
        row_bg = MONTH_FILLS[month_fill_idx]

    rfill = PatternFill("solid", fgColor=row_bg)
    ws_road.row_dimensions[row_n].height = 20

    # Proje badge (config.PROJECT_THEMES'den)
    _theme = cfg.PROJECT_THEMES.get(proj, {})
    proj_color = _theme.get("badge_color", "333333")
    proj_bg    = _theme.get("badge_bg",    "EEEEEE")
    pc = ws_road.cell(row=row_n, column=2, value=proj)
    pc.font = Font(bold=True, size=9, color=proj_color, name="Calibri")
    pc.fill = PatternFill("solid", fgColor=proj_bg)
    pc.alignment = Alignment(horizontal="center", vertical="center")
    pc.border = BORDER2

    # Anahtar
    kc = ws_road.cell(row=row_n, column=3, value=("  └ " if is_sub else "") + item["Anahtar"])
    kc.font = Font(bold=not is_sub, italic=is_sub, color="1F4E79" if not is_sub else "666666", size=9, name="Calibri")
    kc.fill = rfill
    kc.alignment = Alignment(horizontal="left", vertical="center")
    kc.border = BORDER2

    # Görev adı
    ozet = ("    " if is_sub else "") + item["Özet"]
    gc = ws_road.cell(row=row_n, column=4, value=ozet)
    gc.font = Font(bold=not is_sub, italic=is_sub, size=10 if not is_sub else 9,
                   color="111111" if not is_sub else "444444", name="Calibri")
    gc.fill = rfill
    gc.alignment = Alignment(horizontal="left", vertical="center")
    gc.border = BORDER2

    # Durum
    durum_colors = {
        "Yapılacaklar": ("7F6000","FFF2CC"),
        "Devam Ediyor": ("1F4E79","DEEAF1"),
        "Tamamlandı":   ("1E4620","E2EFDA"),
        "Beklemede":    ("7B3400","FCE4D6"),
    }
    df, db = durum_colors.get(status, ("333333","F5F5F5"))
    dc = ws_road.cell(row=row_n, column=5, value=status)
    dc.font = Font(bold=True, size=9, color=df, name="Calibri")
    dc.fill = PatternFill("solid", fgColor=db)
    dc.alignment = Alignment(horizontal="center", vertical="center")
    dc.border = BORDER2

    # Başlangıç Tarihi
    sc = ws_road.cell(row=row_n, column=6, value=item["Başlangıç Tarihi"])
    sc.font = Font(size=10, color="375623", bold=True, name="Calibri")
    sc.fill = rfill
    sc.alignment = Alignment(horizontal="center", vertical="center")
    sc.border = BORDER2

    # Son Tarih
    tc = ws_road.cell(row=row_n, column=7, value=item["Son Tarih"])
    tc.font = Font(size=10, color="1B3A6B", bold=True, name="Calibri")
    tc.fill = rfill
    tc.alignment = Alignment(horizontal="center", vertical="center")
    tc.border = BORDER2

    # Kalan Gün
    if is_done:
        if early_days:
            kalan_val, kalan_color, kalan_bg = f"{early_days} gun ERKEN", "148F3A", "D4EFDF"
        else:
            kalan_val, kalan_color, kalan_bg = "✓ TAMAM", "148F3A", "D4EFDF"
    elif late_start_days:
        kalan_val, kalan_color, kalan_bg = f"BASLANGIC GECTI ({late_start_days} g)", "D35400", "FFE5CC"
    elif remaining is None:
        kalan_val, kalan_color, kalan_bg = "—", "888888", row_bg
    elif remaining < 0:
        kalan_val, kalan_color, kalan_bg = f"Gecti! ({abs(remaining)} gun)", "CC0000", "FFD0D0"
    elif remaining == 0:
        kalan_val, kalan_color, kalan_bg = "BUGUN!", "FF6600", "FFEEDD"
    elif remaining <= 3:
        kalan_val, kalan_color, kalan_bg = f"{remaining} gun", "CC0000", "FFE8E8"
    elif remaining <= 7:
        kalan_val, kalan_color, kalan_bg = f"{remaining} gun", "D35400", "FFF4E0"
    elif remaining <= 14:
        kalan_val, kalan_color, kalan_bg = f"{remaining} gun", "CC8800", "FFFBE6"
    else:
        kalan_val, kalan_color, kalan_bg = f"{remaining} gun", "27AE60", row_bg

    kgc = ws_road.cell(row=row_n, column=8, value=kalan_val)
    kgc.font = Font(bold=True, size=10, color=kalan_color, name="Calibri")
    kgc.fill = PatternFill("solid", fgColor=kalan_bg)
    kgc.alignment = Alignment(horizontal="center", vertical="center")
    kgc.border = BORDER2

    # Zaman Çubuğu
    if is_done:
        if early_days:
            bar_val, bar_color, bar_bg = f"██████████  ✓ {early_days} GUN ERKEN BITIRILDI", "148F3A", "D4EFDF"
        else:
            bar_val, bar_color, bar_bg = "██████████  ✓ TAMAMLANDI", "148F3A", "D4EFDF"
    elif late_start_days:
        bar_val, bar_color, bar_bg = f"⚠ BASLANGIC GECIKTI ({late_start_days} GUN)", "D35400", "FFE5CC"
    elif remaining is None:
        bar_val, bar_color, bar_bg = "Tarih yok", "AAAAAA", "F5F5F5"
    elif remaining < 0:
        bar_val, bar_color, bar_bg = "██████████  SURESI DOLDU", "CC0000", "FFD0D0"
    elif remaining == 0:
        bar_val, bar_color, bar_bg = "██████████  BUGUN SON GUN!", "FF6600", "FFEEDD"
    else:
        MAX = cfg.PROGRESS_WINDOW_ROADMAP
        ratio = min(max(0, MAX - remaining) / MAX, 1.0)
        filled = round(ratio * 10)
        bar = "█" * filled + "░" * (10 - filled)
        if remaining > 30:
            bar_color, bar_bg = "27AE60", "E8F5E9"
        elif remaining > 14:
            bar_color, bar_bg = "2980B9", "EBF5FB"
        elif remaining > 7:
            bar_color, bar_bg = "CC8800", "FFF9E6"
        elif remaining > 3:
            bar_color, bar_bg = "D35400", "FFF0E0"
        else:
            bar_color, bar_bg = "CC0000", "FFE8E8"
        bar_val = f"{bar}  {remaining} gun kaldi"

    zc = ws_road.cell(row=row_n, column=9, value=bar_val)
    zc.font = Font(size=10, color=bar_color, bold=(remaining is not None and remaining <= 14), name="Calibri")
    zc.fill = PatternFill("solid", fgColor=bar_bg)
    zc.alignment = Alignment(horizontal="left", vertical="center")
    zc.border = BORDER2

    # Atanan
    ac = ws_road.cell(row=row_n, column=10, value=item["Atanan"])
    ac.font = Font(size=9, color="444444", name="Calibri")
    ac.fill = rfill
    ac.alignment = Alignment(horizontal="left", vertical="center")
    ac.border = BORDER2

    row_n += 1

ws_road.sheet_view.showGridLines = False
ws_road.sheet_view.zoomScale = 100
ws_road.sheet_properties.tabColor = "2E75B6"

output_path = f"/opt/jira_rapor/raporlar/Digital_Donusum_{datetime.now().strftime('%d.%m.%Y')}.xlsx"
wb.save(output_path)
print(f"\nRapor kaydedildi: {output_path}")

#!/usr/bin/env python3
"""Jira <-> Excel raporu dogrulama scripti.

Akis:
1) Jira'dan TUM issue'lari (tum projeler, paginated) cek -> JSON dump
2) En son Excel raporunu oku, her sheetdeki Anahtar/Durum/Tarih bilgilerini cek
3) Iki kaynagi karsilastir, tutarsizliklari listele
4) Konsola + verify_log.txt'ye rapor yaz

Kullanim:
    python3 verify_report.py
"""
import sys, os, json, glob, base64, urllib.request
from datetime import datetime, date
sys.path.insert(0, "/opt/jira_rapor")
import config as cfg

# ---- Logger ---------------------------------------------------------------
LOG_FILE = "/opt/jira_rapor/verify_log.txt"
_log_lines = []

def log(msg, also_print=True):
    line = msg if isinstance(msg, str) else str(msg)
    _log_lines.append(line)
    if also_print:
        print(line)

def flush_log():
    with open(LOG_FILE, "w", encoding="utf-8") as f:
        f.write("\n".join(_log_lines))

# ---- Jira fetch (paginated) ----------------------------------------------
_creds = base64.b64encode(f"{cfg.JIRA_EMAIL}:{cfg.JIRA_TOKEN}".encode()).decode()
_headers = {
    "Authorization": f"Basic {_creds}",
    "Content-Type": "application/json",
    "Accept": "application/json",
}

def fetch_all(project_key):
    """Sayfalama destekli: tum issue'lari donder."""
    fields = ["summary", "status", "assignee", "priority", "issuetype",
              "created", "duedate", "reporter", "labels",
              "subtasks", "parent", "description", cfg.JIRA_START_DATE_FIELD,
              "resolutiondate"]
    issues = []
    next_token = None
    page_count = 0
    while True:
        body = {
            "jql": f"project={project_key} ORDER BY key ASC",
            "maxResults": 500,
            "fields": fields,
        }
        if next_token:
            body["nextPageToken"] = next_token
        payload = json.dumps(body).encode("utf-8")
        req = urllib.request.Request(
            f"{cfg.JIRA_BASE}/rest/api/3/search/jql",
            data=payload, headers=_headers, method="POST"
        )
        with urllib.request.urlopen(req) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        issues.extend(data.get("issues", []))
        page_count += 1
        if data.get("isLast", True):
            break
        next_token = data.get("nextPageToken")
        if not next_token:
            break
    return issues, page_count

def issue_to_summary(issue):
    """Karsilastirma icin sadelestirilmis veri."""
    f = issue["fields"]
    return {
        "key": issue["key"],
        "summary": f["summary"],
        "status": f["status"]["name"],
        "status_category": (f["status"].get("statusCategory") or {}).get("key", ""),
        "assignee": (f.get("assignee") or {}).get("displayName", ""),
        "priority": (f.get("priority") or {}).get("name", ""),
        "issuetype": f["issuetype"]["name"],
        "is_subtask": bool(f.get("parent") or f["issuetype"].get("subtask")),
        "parent": (f.get("parent") or {}).get("key"),
        "duedate": f.get("duedate"),  # ISO format YYYY-MM-DD or None
        "startdate": f.get(cfg.JIRA_START_DATE_FIELD),
        "labels": f.get("labels") or [],
        "resolutiondate": (f.get("resolutiondate") or "")[:10] if f.get("resolutiondate") else None,
    }

# ---- Excel okuma ----------------------------------------------------------
def normalize_key(raw):
    """'  └ GNDFAB-101' veya '  └ ⚠ GNDFAB-101' -> 'GNDFAB-101'"""
    if not raw:
        return None
    s = str(raw).strip()
    # On-ekleri temizle
    for pre in ["└", "⚠", " "]:
        while s.startswith(pre):
            s = s[1:].strip()
    return s.strip()

def parse_tr_date(s):
    """'07.05.2026' -> date(2026,5,7) | None"""
    if not s or not isinstance(s, str):
        return None
    s = s.strip()
    if len(s) < 10:
        return None
    try:
        return date(int(s[6:10]), int(s[3:5]), int(s[:2]))
    except Exception:
        return None

def read_excel_panos(xlsx_path):
    """Her '* Panosu' sheet'inden anahtar->satir bilgisi cikarir."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    panos = {}  # {sheet_name: {key: row_dict}}

    for sheet_name in wb.sheetnames:
        if not sheet_name.endswith(" Panosu"):
            continue
        ws = wb[sheet_name]
        # Header satiri
        header_row = None
        headers = []
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_col=20, values_only=True), start=1):
            if row and "Anahtar" in [str(c) for c in row if c]:
                header_row = r_idx
                headers = list(row)
                break
        if header_row is None:
            continue

        col_idx = {h: i for i, h in enumerate(headers) if h}
        rows_dict = {}
        for r_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, max_col=20, values_only=True), start=header_row+1):
            if not row or not row[col_idx["Anahtar"]]:
                continue
            key = normalize_key(row[col_idx["Anahtar"]])
            if not key:
                continue
            rows_dict[key] = {
                "row_num": r_idx,
                "summary": row[col_idx.get("Özet", -1)] if "Özet" in col_idx else "",
                "status": row[col_idx.get("Durum", -1)] if "Durum" in col_idx else "",
                "assignee": row[col_idx.get("Atanan", -1)] if "Atanan" in col_idx else "",
                "duedate_str": row[col_idx.get("Son Tarih", -1)] if "Son Tarih" in col_idx else "",
                "startdate_str": row[col_idx.get("Başlangıç Tarihi", -1)] if "Başlangıç Tarihi" in col_idx else "",
            }
        panos[sheet_name] = rows_dict
    return panos

# ---- Karsilastirma --------------------------------------------------------
def compare(jira_issues_by_proj, excel_panos):
    """Her proje icin Jira ve Excel set'lerini kiyasla.
    Donen: dict[proje]: {missing_in_excel, extra_in_excel, status_mismatch, date_mismatch}
    """
    findings = {}

    sheet_for_proj = {p: f"{p} Panosu" for p in jira_issues_by_proj.keys()}

    for proj, issues in jira_issues_by_proj.items():
        sheet_name = sheet_for_proj[proj]
        excel_rows = excel_panos.get(sheet_name, {})

        jira_keys = {iss["key"]: iss for iss in issues}
        excel_keys = set(excel_rows.keys())
        jira_set = set(jira_keys.keys())

        missing_in_excel = sorted(jira_set - excel_keys)
        extra_in_excel = sorted(excel_keys - jira_set)

        status_mismatch = []
        date_mismatch = []

        for key in (jira_set & excel_keys):
            ji = jira_keys[key]
            ei = excel_rows[key]
            # Status karsilastirmasi (esnek: Jira tam adi vs Excel tam adi)
            if str(ji["status"]).strip() != str(ei["status"]).strip():
                status_mismatch.append({
                    "key": key,
                    "jira_status": ji["status"],
                    "excel_status": ei["status"],
                })
            # Tarih karsilastirmasi
            j_due = ji["duedate"]  # YYYY-MM-DD
            e_due = parse_tr_date(ei["duedate_str"])
            j_due_d = None
            if j_due:
                try:
                    j_due_d = date.fromisoformat(j_due[:10])
                except Exception:
                    pass
            if j_due_d != e_due:
                date_mismatch.append({
                    "key": key,
                    "jira_due": j_due,
                    "excel_due": ei["duedate_str"],
                    "field": "duedate",
                })
            # Start date
            j_start = ji["startdate"]
            e_start = parse_tr_date(ei["startdate_str"])
            j_start_d = None
            if j_start:
                try:
                    j_start_d = date.fromisoformat(j_start[:10])
                except Exception:
                    pass
            if j_start_d != e_start:
                date_mismatch.append({
                    "key": key,
                    "jira_start": j_start,
                    "excel_start": ei["startdate_str"],
                    "field": "startdate",
                })

        findings[proj] = {
            "jira_count": len(jira_set),
            "excel_count": len(excel_keys),
            "missing_in_excel": missing_in_excel,
            "extra_in_excel": extra_in_excel,
            "status_mismatch": status_mismatch,
            "date_mismatch": date_mismatch,
        }
    return findings

# ---- Main -----------------------------------------------------------------
def main():
    log("=" * 70)
    log(f"VERIFY RAPOR DOGRULAMA  -  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    log("=" * 70)
    log("")

    # 1) Jira'dan tum projelerin issue'larini cek
    log(">>> 1) JIRA verisi cekiliyor (tum projeler, paginated)...")
    jira_data = {}  # {proj: [issue_dict, ...]}
    jira_summaries = {}  # {proj: [{key,...}]}
    for proj in cfg.PROJECTS:
        log(f"   - {proj} fetch...")
        issues_raw, page_count = fetch_all(proj)
        jira_data[proj] = issues_raw
        jira_summaries[proj] = [issue_to_summary(i) for i in issues_raw]
        log(f"     {len(issues_raw)} issue, {page_count} sayfa")

    # JSON dump
    dump_path = f"/opt/jira_rapor/jira_dump_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    with open(dump_path, "w", encoding="utf-8") as f:
        json.dump({
            "timestamp": datetime.now().isoformat(),
            "summaries": jira_summaries,  # sadelestirilmis (kucuk)
        }, f, ensure_ascii=False, indent=2)
    log(f"   JSON dump: {dump_path}")
    log("")

    # 2) En son Excel raporunu oku
    log(">>> 2) EXCEL raporu okunuyor...")
    files = sorted(glob.glob(cfg.RAPOR_DIR + "/Digital_Donusum_*.xlsx"))
    if not files:
        log("HATA: Excel raporu bulunamadi")
        flush_log()
        sys.exit(1)
    latest = files[-1]
    log(f"   Dosya: {latest}")
    excel_panos = read_excel_panos(latest)
    for sn, rows in excel_panos.items():
        log(f"   - {sn}: {len(rows)} satir")
    log("")

    # 3) Karsilastir
    log(">>> 3) KARSILASTIRMA...")
    log("")
    findings = compare(jira_summaries, excel_panos)

    total_issues = 0
    total_problems = 0
    for proj, f in findings.items():
        log("-" * 60)
        log(f"PROJE: {proj}")
        log(f"  Jira: {f['jira_count']} issue   |   Excel: {f['excel_count']} satir")
        total_issues += f["jira_count"]

        if f["missing_in_excel"]:
            log(f"  [HATA] Excel'de EKSIK ({len(f['missing_in_excel'])}): {', '.join(f['missing_in_excel'])}")
            total_problems += len(f["missing_in_excel"])
        if f["extra_in_excel"]:
            log(f"  [HATA] Excel'de FAZLA ({len(f['extra_in_excel'])}): {', '.join(f['extra_in_excel'])}")
            total_problems += len(f["extra_in_excel"])
        if f["status_mismatch"]:
            log(f"  [UYARI] Durum farki ({len(f['status_mismatch'])}):")
            for m in f["status_mismatch"][:10]:
                log(f"     {m['key']}: Jira='{m['jira_status']}' | Excel='{m['excel_status']}'")
            if len(f["status_mismatch"]) > 10:
                log(f"     ... ve {len(f['status_mismatch']) - 10} tane daha")
            total_problems += len(f["status_mismatch"])
        if f["date_mismatch"]:
            log(f"  [UYARI] Tarih farki ({len(f['date_mismatch'])}):")
            for m in f["date_mismatch"][:10]:
                log(f"     {m['key']} ({m['field']}): Jira={m.get('jira_due') or m.get('jira_start')} | Excel={m.get('excel_due') or m.get('excel_start')}")
            if len(f["date_mismatch"]) > 10:
                log(f"     ... ve {len(f['date_mismatch']) - 10} tane daha")
            total_problems += len(f["date_mismatch"])

        if not (f["missing_in_excel"] or f["extra_in_excel"] or f["status_mismatch"] or f["date_mismatch"]):
            log("  [OK] Hicbir tutarsizlik yok")

    log("")
    log("=" * 70)
    log(f"OZET: {total_issues} Jira issue   |   {total_problems} tutarsizlik")
    log("=" * 70)

    flush_log()
    return total_problems

if __name__ == "__main__":
    sys.exit(main())
